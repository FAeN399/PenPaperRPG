#!/usr/bin/env python3
"""
Dynamic Pathfinder 2e Character Sheet - Excel Generator
Uses FORMULAS for all calculations (not hardcoded values)
Makes the spreadsheet dynamic - change an ability score and everything updates!
"""

import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers

# Color coding per XLSX skill standards
BLUE_INPUT = Font(name="Arial", size=11, color="0000FF")  # Hardcoded inputs
BLACK_FORMULA = Font(name="Arial", size=11, color="000000")  # Formulas
GOLD_EMPHASIS = Font(name="Arial", size=12, bold=True, color="DAA520")
HEADING_FONT = Font(name="Arial", size=12, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Arial", size=11, bold=True)
NORMAL_FONT = Font(name="Arial", size=11)

GOLD_FILL = PatternFill(start_color="DAA520", end_color="DAA520", fill_type="solid")
DARK_FILL = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def get_name(entity_id):
    if not entity_id:
        return "—"
    return entity_id.split(".")[-1].replace("-", " ").title()

def get_prof_rank_value(rank):
    """Convert rank to numeric value for formula"""
    ranks = {"untrained": 0, "trained": 2, "expert": 4, "master": 6, "legendary": 8}
    return ranks.get(rank, 0)

def create_character_excel(json_path, output_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    char = data['character']
    meta = char.get('metadata', {})
    identity = char.get('identity', {})
    abilities = char.get('abilityScores', {}).get('final', {})
    prof = char.get('proficiencies', {})
    derived = char.get('derived', {})

    wb = Workbook()
    wb.remove(wb.active)

    # ========== SHEET 1: CHARACTER OVERVIEW ==========
    ws = wb.create_sheet("Character")

    # Title
    ws['A1'] = meta.get('name', 'Character')
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(name="Arial", size=20, bold=True, color="DAA520")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Identity Section
    row = 3
    ws[f'A{row}'] = 'CHARACTER IDENTITY'
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'].font = HEADING_FONT
    ws[f'A{row}'].fill = DARK_FILL
    ws[f'A{row}'].alignment = Alignment(horizontal="center")
    row += 1

    identity_data = [
        ['Level', identity.get('level', 1)],
        ['Player', meta.get('player', '—')],
        ['Campaign', meta.get('campaign', '—')],
        ['Ancestry', get_name(identity.get('ancestryId'))],
        ['Heritage', get_name(identity.get('heritageId'))],
        ['Background', get_name(identity.get('backgroundId'))],
        ['Class', get_name(identity.get('classId'))],
        ['Deity', get_name(identity.get('deityId'))],
        ['Alignment', identity.get('alignment', '—')],
    ]

    for label, value in identity_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = BOLD_FONT
        ws[f'A{row}'].fill = LIGHT_FILL
        ws[f'A{row}'].border = THIN_BORDER
        ws[f'A{row}'].alignment = Alignment(horizontal="right")
        ws[f'B{row}'].font = BLUE_INPUT if label == 'Level' else NORMAL_FONT
        ws[f'B{row}'].border = THIN_BORDER
        row += 1

    # Store level cell for formulas
    level_cell = 'B4'  # Level is at row 4
    row += 1

    # ========== ABILITY SCORES WITH FORMULAS ==========
    ws[f'A{row}'] = 'ABILITY SCORES'
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].font = HEADING_FONT
    ws[f'A{row}'].fill = DARK_FILL
    ws[f'A{row}'].alignment = Alignment(horizontal="center")
    row += 1

    # Ability headers
    ability_list = ['STR', 'DEX', 'CON', 'INT', 'WIS', 'CHA']
    for col, ability in enumerate(ability_list, start=1):
        cell = ws.cell(row, col)
        cell.value = ability
        cell.font = HEADING_FONT
        cell.fill = DARK_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    row += 1
    ability_score_row = row

    # Ability scores (BLUE = input values)
    for col, ability in enumerate(ability_list, start=1):
        cell = ws.cell(row, col)
        cell.value = abilities.get(ability, 10)
        cell.font = Font(name="Arial", size=14, bold=True, color="0000FF")  # BLUE INPUT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = '0'

    row += 1

    # Ability modifiers (BLACK = formulas)
    for col, ability in enumerate(ability_list, start=1):
        cell = ws.cell(row, col)
        # FORMULA: =(score - 10) / 2
        score_cell = ws.cell(ability_score_row, col).coordinate
        cell.value = f'=INT(({score_cell}-10)/2)'
        cell.font = BLACK_FORMULA
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = '0'  # Show as integer

    row += 2

    # Store ability modifier cells for later formulas
    ability_mod_row = row - 1
    ability_cells = {ability: ws.cell(ability_score_row, col).coordinate for col, ability in enumerate(ability_list, start=1)}
    mod_cells = {ability: ws.cell(ability_mod_row, col).coordinate for col, ability in enumerate(ability_list, start=1)}

    # ========== CORE STATISTICS ==========
    ws[f'A{row}'] = 'CORE STATISTICS'
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'].font = HEADING_FONT
    ws[f'A{row}'].fill = DARK_FILL
    ws[f'A{row}'].alignment = Alignment(horizontal="center")
    row += 1

    hp = derived.get('hitPoints', {})

    core_stats = [
        ['Hit Points', f"{hp.get('current', 0)}/{hp.get('max', 0)}"],
        ['Armor Class', derived.get('armorClass', 10)],
        ['Class DC', derived.get('classDC', 10)],
        ['Speed', f"{derived.get('speeds', {}).get('land', 25)} ft"],
    ]

    for label, value in core_stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = BOLD_FONT
        ws[f'A{row}'].fill = LIGHT_FILL
        ws[f'A{row}'].border = THIN_BORDER
        ws[f'A{row}'].alignment = Alignment(horizontal="right")
        ws[f'B{row}'].font = GOLD_EMPHASIS
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'B{row}'].alignment = Alignment(horizontal="center")
        row += 1

    # ========== SHEET 2: SKILLS & SAVES ==========
    ws_skills = wb.create_sheet("Skills & Saves")
    row = 1

    # SAVING THROWS
    ws_skills[f'A{row}'] = 'SAVING THROWS'
    ws_skills.merge_cells(f'A{row}:D{row}')
    ws_skills[f'A{row}'].font = HEADING_FONT
    ws_skills[f'A{row}'].fill = DARK_FILL
    ws_skills[f'A{row}'].alignment = Alignment(horizontal="center")
    row += 1

    ws_skills[f'A{row}'] = 'Save'
    ws_skills[f'B{row}'] = 'Ability'
    ws_skills[f'C{row}'] = 'Rank'
    ws_skills[f'D{row}'] = 'Modifier'
    for col in range(1, 5):
        cell = ws_skills.cell(row, col)
        cell.font = HEADING_FONT
        cell.fill = DARK_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Save formulas
    saves_data = [
        ('Fortitude', 'CON', prof.get('saves', {}).get('fortitude', 'untrained')),
        ('Reflex', 'DEX', prof.get('saves', {}).get('reflex', 'untrained')),
        ('Will', 'WIS', prof.get('saves', {}).get('will', 'untrained')),
    ]

    for save_name, ability, rank in saves_data:
        ws_skills[f'A{row}'] = save_name
        ws_skills[f'B{row}'] = ability
        ws_skills[f'C{row}'] = rank.title()

        # FORMULA: = ability_mod + (rank_value + level)
        rank_val = get_prof_rank_value(rank)
        # Reference the Character sheet for level and ability mod
        ws_skills[f'D{row}'] = f'=Character!{mod_cells[ability]} + (Character!{level_cell} + {rank_val})'

        ws_skills[f'A{row}'].border = THIN_BORDER
        ws_skills[f'B{row}'].border = THIN_BORDER
        ws_skills[f'C{row}'].border = THIN_BORDER
        ws_skills[f'D{row}'].font = BLACK_FORMULA
        ws_skills[f'D{row}'].border = THIN_BORDER
        ws_skills[f'D{row}'].alignment = Alignment(horizontal="center")
        ws_skills[f'D{row}'].number_format = '0'
        row += 1

    row += 1

    # SKILLS
    ws_skills[f'A{row}'] = 'SKILLS'
    ws_skills.merge_cells(f'A{row}:D{row}')
    ws_skills[f'A{row}'].font = HEADING_FONT
    ws_skills[f'A{row}'].fill = DARK_FILL
    ws_skills[f'A{row}'].alignment = Alignment(horizontal="center")
    row += 1

    ws_skills[f'A{row}'] = 'Skill'
    ws_skills[f'B{row}'] = 'Ability'
    ws_skills[f'C{row}'] = 'Rank'
    ws_skills[f'D{row}'] = 'Modifier'
    for col in range(1, 5):
        cell = ws_skills.cell(row, col)
        cell.font = HEADING_FONT
        cell.fill = DARK_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1

    skill_ability_map = {
        'Acrobatics': 'DEX', 'Arcana': 'INT', 'Athletics': 'STR',
        'Crafting': 'INT', 'Deception': 'CHA', 'Diplomacy': 'CHA',
        'Intimidation': 'CHA', 'Medicine': 'WIS', 'Nature': 'WIS',
        'Occultism': 'INT', 'Performance': 'CHA', 'Religion': 'WIS',
        'Society': 'INT', 'Stealth': 'DEX', 'Survival': 'WIS',
        'Thievery': 'DEX'
    }

    skills_prof = prof.get('skills', {})
    for skill_name in sorted(skills_prof.keys()):
        rank = skills_prof[skill_name]
        ability = skill_ability_map.get(skill_name, 'INT')
        rank_val = get_prof_rank_value(rank)

        ws_skills[f'A{row}'] = skill_name
        ws_skills[f'B{row}'] = ability
        ws_skills[f'C{row}'] = rank.title()

        # FORMULA: = ability_mod + (rank_value + level)
        ws_skills[f'D{row}'] = f'=Character!{mod_cells[ability]} + (Character!{level_cell} + {rank_val})'

        ws_skills[f'A{row}'].border = THIN_BORDER
        ws_skills[f'B{row}'].border = THIN_BORDER
        ws_skills[f'C{row}'].border = THIN_BORDER
        ws_skills[f'D{row}'].font = BLACK_FORMULA
        ws_skills[f'D{row}'].border = THIN_BORDER
        ws_skills[f'D{row}'].alignment = Alignment(horizontal="center")
        ws_skills[f'D{row}'].number_format = '0'
        row += 1

    # ========== SHEET 3: FEATS ==========
    ws_feats = wb.create_sheet("Feats")
    row = 1

    ws_feats[f'A{row}'] = 'FEATS & FEATURES'
    ws_feats.merge_cells(f'A{row}:C{row}')
    ws_feats[f'A{row}'].font = HEADING_FONT
    ws_feats[f'A{row}'].fill = DARK_FILL
    ws_feats[f'A{row}'].alignment = Alignment(horizontal="center")
    row += 1

    ws_feats[f'A{row}'] = 'Feat Name'
    ws_feats[f'B{row}'] = 'Level'
    ws_feats[f'C{row}'] = 'Source'
    for col in range(1, 4):
        cell = ws_feats.cell(row, col)
        cell.font = HEADING_FONT
        cell.fill = DARK_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1

    feats = char.get('feats', [])
    for feat in sorted(feats, key=lambda f: f.get('level', 1)):
        ws_feats[f'A{row}'] = get_name(feat.get('id', ''))
        ws_feats[f'B{row}'] = feat.get('level', 1)
        ws_feats[f'C{row}'] = feat.get('grantedBy', '').title()

        ws_feats[f'A{row}'].border = THIN_BORDER
        ws_feats[f'B{row}'].border = THIN_BORDER
        ws_feats[f'C{row}'].border = THIN_BORDER
        ws_feats[f'B{row}'].alignment = Alignment(horizontal="center")
        row += 1

    # ========== SHEET 4: EQUIPMENT ==========
    ws_equip = wb.create_sheet("Equipment")
    row = 1

    ws_equip[f'A{row}'] = 'EQUIPMENT & INVENTORY'
    ws_equip.merge_cells(f'A{row}:D{row}')
    ws_equip[f'A{row}'].font = HEADING_FONT
    ws_equip[f'A{row}'].fill = DARK_FILL
    ws_equip[f'A{row}'].alignment = Alignment(horizontal="center")
    row += 1

    ws_equip[f'A{row}'] = 'Item'
    ws_equip[f'B{row}'] = 'Qty'
    ws_equip[f'C{row}'] = 'Bulk'
    ws_equip[f'D{row}'] = 'Notes'
    for col in range(1, 5):
        cell = ws_equip.cell(row, col)
        cell.font = HEADING_FONT
        cell.fill = DARK_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1

    equipment = char.get('equipment', [])
    for item in equipment:
        ws_equip[f'A{row}'] = item.get('name', 'Unknown')
        ws_equip[f'B{row}'] = item.get('quantity', 1)
        ws_equip[f'C{row}'] = str(item.get('bulk', '—'))
        ws_equip[f'D{row}'] = item.get('notes', '')

        for col in range(1, 5):
            ws_equip.cell(row, col).border = THIN_BORDER
        ws_equip[f'B{row}'].alignment = Alignment(horizontal="center")
        ws_equip[f'C{row}'].alignment = Alignment(horizontal="center")
        row += 1

    # ========== SHEET 5: NOTES ==========
    notes = char.get('notes', {})
    if any(notes.values()):
        ws_notes = wb.create_sheet("Notes")
        row = 1

        ws_notes[f'A{row}'] = 'CHARACTER DETAILS'
        ws_notes.merge_cells(f'A{row}:D{row}')
        ws_notes[f'A{row}'].font = HEADING_FONT
        ws_notes[f'A{row}'].fill = DARK_FILL
        ws_notes[f'A{row}'].alignment = Alignment(horizontal="center")
        row += 2

        for section, content in [
            ('Appearance', notes.get('appearance')),
            ('Backstory', notes.get('backstory')),
            ('Allies', notes.get('allies')),
            ('Campaign Notes', notes.get('campaigns'))
        ]:
            if content:
                ws_notes[f'A{row}'] = section
                ws_notes[f'A{row}'].font = BOLD_FONT
                row += 1
                ws_notes[f'A{row}'] = content
                ws_notes.merge_cells(f'A{row}:D{row}')
                ws_notes[f'A{row}'].alignment = Alignment(wrap_text=True, vertical="top")
                ws_notes.row_dimensions[row].height = 60
                row += 2

    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws_skills.column_dimensions['A'].width = 18
    ws_skills.column_dimensions['B'].width = 10
    ws_skills.column_dimensions['C'].width = 15
    ws_skills.column_dimensions['D'].width = 12
    ws_feats.column_dimensions['A'].width = 30
    ws_feats.column_dimensions['B'].width = 8
    ws_feats.column_dimensions['C'].width = 15
    ws_equip.column_dimensions['A'].width = 25
    ws_equip.column_dimensions['D'].width = 30

    wb.save(output_path)
    print(f"Dynamic Excel character sheet created: {output_path}")
    print("IMPORTANT: Run recalc.py to calculate formulas!")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python create-excel-character-sheet.py <character.json> <output.xlsx>")
        sys.exit(1)

    create_character_excel(sys.argv[1], sys.argv[2])
